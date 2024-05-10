import openpyxl as xl
from openpyxl.chart import BarChart, Reference


# An employee of the company mistakenly forgets to put the 10% discount on the price. However, he has already saved
# millions of data without putting that 10% discount. So, this program helps to automate the cell data and save the
# data in the Excel spreadsheet file of the millions of data without need to change them manually from excel worksheet.


def process_workbook(filename):
    workbook = xl.load_workbook(filename)
    sheet = workbook["Sheet1"]
    for row in range(3, sheet.max_row + 1):
        # Applying 10% discount to the column 6
        corrected_val = sheet.cell(row, 6).value * 0.9
        sheet.cell(row, 6).value = corrected_val

    # Add new chart
    values = Reference(sheet,
                       min_row=3,
                       max_row=sheet.max_row,
                       min_col=6,
                       max_col=6)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, "G5")
    workbook.save(filename)


process_workbook("Book1.xlsx")
